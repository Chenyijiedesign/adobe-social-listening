#!/usr/bin/env python3
"""
Prepare LinkedIn visualization data from tagged Excel file.
Output: linkedin_data.json  (sits alongside viz_data.json for index.html)
"""

import json, re, sys, os
import numpy as np
from collections import Counter

try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.run([sys.executable,'-m','pip','install','openpyxl','-q'])
    import openpyxl

INPUT  = os.path.join(os.path.dirname(__file__),
         "../linkedin/hashtag_results_merged.xlsx")
OUTPUT = "linkedin_data.json"

# ── Tag normalization ─────────────────────────────────────────────────────────

EMOTION_NORM = {
    "inspiration": "excitement",
    "joy":         "excitement",
    "frustration": "resentment",
    "critique":    "skepticism",
}
AI_REL_NORM = {
    "critique":    "none",
    "integration": "adaptation",
    "comparison":  "none",
    "coexistence": "adaptation",
}
ACTIVITY_NORM = {
    "neutral":             "discussion",
    "recommendation":      "workflow_share",
    "educational_content": "tutorial",
}

def get_category(query):
    q = query.lower().lstrip('#')
    # UX must be checked before Design — 'design' is a substring of 'uxdesign', 'figmadesign', 'webdesign'
    if q in ('uxdesign','figmadesign','webdesign','frontend','userexperience'):
        return "UX & Product"
    if any(k in q for k in ['adobe','photoshop','aftereffects','lightroom',
                             'adobeanimate','premierepro','adobeillustrator','adobefirefly']):
        return "Adobe"
    if any(k in q for k in ['stablediffusion','midjourney','comfyui','aiart',
                             'chatgpt','claudeai']):
        return "AI & Visual"
    if any(k in q for k in ['graphicdesign','logodesign','design','typography','branding']):
        return "Design & Branding"
    if any(k in q for k in ['blender','gamedev','3dmodeling','unrealengine','unity3d']):
        return "3D & Gaming"
    if any(k in q for k in ['vfx','filmmakers','cinematography','editors']):
        return "Motion & VFX"
    return "Other"

def clean_text(s):
    s = re.sub(r'http\S+', '', s or '')
    s = re.sub(r'hashtag\s*\n', '', s)   # remove "hashtag\n#tag" patterns
    s = re.sub(r'#\w+', '', s)            # remove hashtags
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def norm_bool(v):
    if isinstance(v, bool): return v
    return str(v).strip().upper() in ("TRUE", "1", "YES")

# ── Load Excel ────────────────────────────────────────────────────────────────

print("Loading Excel…")
wb = openpyxl.load_workbook(INPUT, read_only=True)
ws = wb['Sheet1']
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
raw = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
print(f"  {len(raw)} rows")

# ── Build posts ───────────────────────────────────────────────────────────────

posts = []
for i, r in enumerate(raw):
    text_full = r.get('text') or ''
    text_clean = clean_text(text_full)

    emotion  = str(r.get('Emotion') or '').strip().lower()
    emotion  = EMOTION_NORM.get(emotion, emotion) or 'neutral'

    activity = str(r.get('Activity') or '').strip().lower()
    activity = ACTIVITY_NORM.get(activity, activity) or 'discussion'

    ai_rel   = str(r.get('AI Relation') or '').strip().lower()
    ai_rel   = AI_REL_NORM.get(ai_rel, ai_rel) or 'none'

    query    = str(r.get('query') or '').strip()
    score    = int(r.get('reactions') or 0)

    # Use first sentence / 120 chars of cleaned text as the "title"
    title_src = text_clean[:300]
    first_line = title_src.split('\n')[0].strip()
    title = first_line[:120] + ('…' if len(first_line) > 120 else '')

    posts.append({
        "id":           f"li-{i}",
        "title":        title,
        "full_text":    text_clean,       # full cleaned text for modal
        "body_preview": text_clean[:180],
        "subreddit":    query,           # keeps same field name as Reddit
        "category":     get_category(query),
        "emotion":      emotion,
        "activity":     activity,
        "ai_related":   norm_bool(r.get('AI Related')),
        "ai_relation":  ai_rel,
        "adobe_related": norm_bool(r.get('Adobe Related')),
        "score":        score,
        "url":          "",              # LinkedIn doesn't expose post URLs via scraping
        "created_at":   str(r.get('scraped_at') or '')[:10],
    })

# ── UMAP ──────────────────────────────────────────────────────────────────────

print("\nComputing UMAP…")
try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    import umap as umap_lib

    corpus = [f"{p['title']} {p['body_preview']}" for p in posts]

    print("  TF-IDF…")
    vec = TfidfVectorizer(max_features=8000, ngram_range=(1,2),
                          sublinear_tf=True, stop_words="english", min_df=2)
    X = vec.fit_transform(corpus)

    print(f"  Matrix {X.shape} — UMAP (≈1-2 min)…")
    reducer = umap_lib.UMAP(n_neighbors=15, min_dist=0.08, n_components=2,
                             metric="cosine", random_state=42, verbose=False)
    coords = reducer.fit_transform(X)

    for i, p in enumerate(posts):
        p["ux"] = round(float(coords[i,0]), 4)
        p["uy"] = round(float(coords[i,1]), 4)
    has_umap = True
    print("  ✓ UMAP done")

except ImportError as e:
    print(f"  ⚠ Skipping ({e})")
    for p in posts:
        p["ux"] = 0.0; p["uy"] = 0.0
    has_umap = False

# ── Subreddit (query) metadata ────────────────────────────────────────────────

sub_info = {}
for p in posts:
    s = p["subreddit"]
    if s not in sub_info:
        sub_info[s] = {"name": s, "category": p["category"], "count": 0}
    sub_info[s]["count"] += 1

subreddits = sorted(sub_info.values(), key=lambda x: (x["category"], -x["count"]))

# ── Save ──────────────────────────────────────────────────────────────────────

out = {"posts": posts, "subreddits": subreddits,
       "has_umap": has_umap, "total": len(posts)}

print(f"\nSaving {OUTPUT}…")
with open(OUTPUT, "w", encoding="utf-8") as f:
    json.dump(out, f, separators=(",",":"))

size_kb = len(json.dumps(out)) / 1024
print(f"✅  {OUTPUT}  ({size_kb:.0f} KB,  {len(posts)} posts)")
